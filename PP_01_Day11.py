#In excel with Huge data
#Writing data
from openpyxl import Workbook, load_workbook

# Create a new workbook
wb = Workbook()
ws = wb.active

# Example: Writing 1,000 rows
for i in range(1, 1001):
    ws.append([f"Row {i}", f"Value {i}"])

# Save the workbook
wb.save("D:\Ankita\Training\Practice.xlsx")

# Load the workbook
wb = load_workbook(r"D:\Ankita\Training\Practice.xlsx", read_only=True)  # Use read_only mode for large files
ws = wb.active

# Read rows one by one
for row in ws.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
    print(row)  # Print the first 10 rows


# Read specific range
for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
    print(row)
